import datetime, csv
import requests
from bs4 import BeautifulSoup
import os
import crolling as crolling
import pymysql
import openpyxl
from selenium import webdriver
import getpass

def make_dir(dir):
    if not os.path.exists(dir):
        os.makedirs(dir)
    return dir

pagePrintCnt = 50
prj_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


user = getpass.getuser()

down_path = 'C:\\Users\\'+user+'\\Downloads'
common_path = prj_path+'/common'
data_path = prj_path+'/data'

keyword_r_path = make_dir(data_path + '/keywordR')

keyword_pc_path = make_dir(data_path + '/keywordPC')
keyword_mb_path = make_dir(data_path + '/keywordMB')

keyword_pcd_path = make_dir(data_path + '/keywordPCD')
keyword_mbd_path = make_dir(data_path + '/keywordMBD')

key_value_file = common_path + '/' + 'key_value.py'
if not os.path.exists(key_value_file):
    with open(key_value_file, 'w') as f:
        f.write("host='0.0.0.0'\n")
        f.write("port=3307\n")
        f.write("user='hee'\n")
        f.write("password=''\n")
        f.write("db=''\n")
        f.write("charset='utf8mb4'\n")
        f.write("nv_ad_id = ''\n")
        f.write("nv_ad_pw = ''\n")
        f.write("du_ad_id = ''\n")
        f.write("du_ad_pw = ''\n")

from common import key_value as key_value

def println(strTxt, cnt):
    if len(strTxt) > cnt:
        cnt = len(strTxt)+1
    return (strTxt + " " * cnt)[:cnt]

def get_strArr(strArr, strArrKey):
    for strStr in strArrKey:
        strArr.append(strStr)
    return strArr

def get_strArr_sub(strArr, strArrKey, strArrKeySub):
    for strR in strArrKey:
        for strRS in strArrKeySub:
            strArr.append(strR + strRS)
    return strArr

def get_key_nv_list():
    for keyJson in crolling.keyNv:
        pStr = println('Key:' + keyJson['key'], 30)
        pStr += println('Name:' + keyJson['name'], 40)
        pStr += println('Mid1:' + keyJson['mid1'], 30)
        pStr += println('Mid2:' + keyJson['mid2'], 30)
        pStr += keyJson['cnt']
        print(pStr)
    print('')

    findKeyArr = []  # 검색 키워드
    for keyJson in crolling.keyNv:
        if keyJson['key'] in findKeyArr:
            None
        else:
            findKeyArr.append(keyJson['key'])

        if keyJson['mid2'] == '':
            keyJson['mid2'] = '사용안함'
        print('검색명=' + keyJson['key'] + '상품명=' + keyJson['mid1'] + '구매처=' + keyJson['mid2'] + '페이지=' + keyJson['cnt'])
    print('')
    for key in crolling.keyList:
        if key in findKeyArr:
            None
        else:
            findKeyArr.append(key)
    if crolling.nv_flag is False:
        return crolling.keyList
    return findKeyArr

def get_soup(strTxt, pg):
    url = 'https://search.shopping.naver.com/search/all.nhn?origQuery=' + strTxt
    url += '&pagingIndex=' + str(pg) + '&pagingSize=40&viewType=list&sort=rel&frm=NVSHPAG&query=' + strTxt
    html = requests.get(url).text
    soup = BeautifulSoup(html, 'lxml') # 'html.parser''lxml'
    return soup

def get_soup_pwlinkMainPC(strTxt):
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'}
    url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + strTxt
    html = requests.get(url, headers=headers).text
    soup = BeautifulSoup(html, 'lxml')
    return soup

def get_soup_pwlinkSub(strTxt, pg):
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'}
    url = 'https://ad.search.naver.com/search.naver?where=ad&query=' + strTxt+'&pagingIndex='+str(pg)
    html = requests.get(url, headers=headers).text
    soup = BeautifulSoup(html, 'lxml')
    return soup

def get_soup_pwlinkMainMB(strTxt):
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'}
    url = 'https://m.ad.search.naver.com/search.naver?where=m_expd&query=' + strTxt
    html = requests.get(url, headers=headers).text
    soup = BeautifulSoup(html, 'lxml')
    return soup

def get_rank_keywordR():
    rankData = {}
    rank_list = sorted(os.listdir(keyword_r_path), reverse=True)
    for dir in rank_list:
        file = keyword_r_path + '/' + dir
        if file.find('xlsx') > 0 and file.find('연관키워드') > 0:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for row in ws.rows:
                name = row[0].value
                pc = row[1].value
                mb = row[2].value
                if name in rankData:
                    continue
                rankData[name] = {'pc': pc, 'mb': mb}
    return rankData

def get_rank_keyword(pc_mb):
    rankData = {}
    if pc_mb == 'pc':
        keyword_path = keyword_pc_path
    else:
        keyword_path = keyword_mb_path
    rank_list = sorted(os.listdir(keyword_path), reverse=True)
    for dir in rank_list:
        file = keyword_path + '/' + dir
        if file.find('xlsx') > 0 and file.find('키워드 목록') > 0 and file.find('~$키워드 목록') == -1:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for row in ws.rows:
                id = row[0].value
                status = row[1].value
                find_key = row[2].value
                cost = row[4].value
                view = row[6].value
                click = row[7].value
                total_cost = row[11].value
                if id == '' or id == '키워드 ID' or find_key in rankData:
                    continue
                if status != '노출가능':
                    continue
                rankData[find_key] = {'status': status, 'cost': cost, 'view': view, 'click': click, 'total_cost': total_cost, 'file_name':file}
    return rankData

def get_rank_keyword_daum(pc_mb):
    rankData = {}
    if pc_mb == 'pcd':
        keyword_path = keyword_pcd_path
    else:
        keyword_path = keyword_mbd_path
    rank_list = sorted(os.listdir(keyword_path), reverse=True)
    for dir in rank_list:
        file = keyword_path + '/' + dir
        if file.find('csv') > 0 and file.find('현황보고서') > 0 and file.find('~$현황보고서') == -1:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for row in ws.rows:
                id = row[0].value
                status = row[1].value
                find_key = row[2].value
                cost = row[4].value
                view = row[6].value
                click = row[7].value
                total_cost = row[11].value
                if id == '' or id == '키워드 ID' or find_key in rankData:
                    continue
                if status != '노출가능':
                    continue
                rankData[find_key] = {'status': status, 'cost': cost, 'view': view, 'click': click, 'total_cost': total_cost, 'file_name':file}
    return rankData

def print_find_text(strKey):
    pStr1 = '============================================================================================'
    pStr2 = str(strKey)
    pStr3 = '--------------------------------------------------------------------------------------------'
    print(pStr1)
    print(pStr2)
    print(pStr3)

def get_keyword_list(pc_mb = None):
    rankKey = {}
    rankKeyR = {}
    # Open database connection
    try:
        db = pymysql.connect(host=key_value.host
                             , port=key_value.port
                             , user=key_value.user
                             , passwd=key_value.password
                             , db=key_value.db
                             , charset=key_value.charset
                             , autocommit=True)

        curs = db.cursor()
        sql = "select * from flybeach.keyword "
        if pc_mb is not None:
            sql += "where pc_mb = '"+pc_mb+"' "
        sql += "order by pc_mb desc, view desc "
        curs.execute(sql)
        rowsKey = curs.fetchall()
        for rows in rowsKey:
            id = rows[0].upper() + ':' + rows[1]
            rankKey[id] = {'view': rows[2], 'click': rows[3], 'cost': rows[4], 'total_cost': rows[5]}

        curs = db.cursor()
        sql = "select * from flybeach.keywordR "
        curs.execute(sql)
        rowsKeyR = curs.fetchall()
        for rows in rowsKeyR:
            rankKeyR[rows[0]] = {'pc_cnt': rows[1], 'mb_cnt': rows[2]}
        return db, rowsKey, rowsKeyR, rankKey, rankKeyR
    except:
        print('db connection error............................................')
        return None, None, None, None, None

def insert_history(db, input_data):
    if db is None:
        return
    # try:
    title = input_data['title'].replace("'", "")
    find_key = input_data['find_key']
    ad = input_data['ad']
    site = input_data['site']
    item = input_data['item']
    page = input_data['page']
    idx = input_data['idx']
    mid1 = input_data['mid1']
    url = input_data['url']
    now = datetime.date.today()
    upStr = now.strftime("%Y-%m-%d")

    nowTime = datetime.datetime.now()
    nowTimeStr = nowTime.strftime("%Y-%m-%d %H:%M:%S")
    # nowDate = datetime.datetime(2009, 5, 5)
    curs = db.cursor()
    sql = "select * from flybeach.history "
    sql += "where update_date=%s and find_key=%s and mid1=%s and ad=%s "
    curs.execute(sql, (upStr, find_key, mid1, ad))
    rows = curs.fetchall()

    if len(rows) > 0:
        #Update
        sql = "update flybeach.history "
        sql += "set title='" + title + "', page=" + str(page) + ", idx=" + str(idx) + ", last_update_date='" + nowTimeStr + "' "
        sql += "where update_date=%s and find_key=%s and mid1=%s "
        curs.execute(sql, (upStr, find_key, mid1))
    else:
        # Insert
        sql = "insert into flybeach.history(title, find_key, ad, site, item, page, idx, mid1, url, update_date) "
        sql += "values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        curs.execute(sql, (title, find_key, ad, site, item, page, idx, mid1, url, upStr))

    db.commit()
    # except Exception as e:
    #     print(e)

def insert_history_pwlink(db, input_data):
    if db is None:
        return
    # try:
    find_key = input_data['find_key']
    main_sub = input_data['main_sub']
    idx = input_data['idx']
    idx_total = input_data['idx_total']
    item = input_data['item']
    item_desc = input_data['item_desc']
    pc_mb = input_data['pc_mb']
    view = input_data['view']
    click = input_data['click']
    cost = input_data['cost']
    total_cost = input_data['total_cost']

    now = datetime.date.today()
    upStr = now.strftime("%Y-%m-%d")

    nowTime = datetime.datetime.now()
    nowTimeStr = nowTime.strftime("%Y-%m-%d %H:%M:%S")
    # nowDate = datetime.datetime(2009, 5, 5)
    curs = db.cursor()
    sql = "select * from flybeach.history_pwlink "
    sql += "where update_date=%s and find_key=%s and pc_mb=%s "
    curs.execute(sql, (upStr, find_key, pc_mb))
    rows = curs.fetchall()

    if len(rows) > 0:
        #Update
        sql = "update flybeach.history_pwlink "
        sql += "set main_sub='" + main_sub + "', idx=" + str(idx) + ", idx_total=" + str(idx_total) + ", last_update_date='" + nowTimeStr + "' "
        sql += ", view=" + str(view) + ", click=" + str(click) + ", cost=" + str(cost) + ", total_cost=" + str(total_cost) + " "
        sql += "where update_date=%s and find_key=%s "
        curs.execute(sql, (upStr, find_key))
    else:
        # Insert
        sql = "insert into flybeach.history_pwlink(find_key, main_sub, idx, idx_total, item, item_desc, update_date, pc_mb, view, click, cost, total_cost) "
        sql += "values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        curs.execute(sql, (find_key, main_sub, idx, idx_total, item, item_desc, upStr, pc_mb, view, click, cost, total_cost))

    db.commit()
    # except Exception as e:
    #     print(e)

def find_page(strTxtPrint, findKey, pg, db):
    idx = 0
    soup = get_soup(findKey, pg)
    for tag in soup.select('li'):
        tagImg = tag.find(class_='img')
        if tagImg:
            idx += 1
            ct = tagImg.contents[1]
            imgName = ct.get('alt')
            pageurl = tagImg.attrs['href']
            fbIdx = tag.text.find('FLYBEACH')
            fbsIdxH = tag.text.find('플라이비치')
            adIdx = str(tag).find('ad _itemSection')
            site = None
            if fbIdx > -1:
                site = 'FLYBEACH'
                psite = 'FLYBEACH    '
            elif fbsIdxH > -1:
                site = '플라이비치'
                psite = '플라이비치  '

            ad = ''
            if site is not None and adIdx > 0:
                ad = '(광고)'

            if site is not None:
                pStr = ad
                pStr += psite
                pStr += println(str(imgName),25)
                pStr += 'Page:' + println(str(pg), 5)
                pStr += 'INDEX:' + println(str(idx), 5)
                pStr += 'MID:' + println(tag.attrs['data-nv-mid'], 20)
                pStr += '      ' + pageurl
                input_data = {'title':strTxtPrint, 'find_key':findKey, 'ad':ad, 'site':site, 'item':str(imgName), 'page':pg, 'idx':idx
                                , 'mid1':tag.attrs['data-nv-mid'], 'url':pageurl}
                insert_history(db, input_data)
                print(pStr)

    if pg % pagePrintCnt == 0:
        print('Process Page:' + str(pg))

def set_make_title(rankKey, rankKeyR, findKey):
    rd = ''
    if rankKeyR is not None and findKey in rankKeyR.keys():
        rd = str(rankKeyR[findKey])

    pcFindKey = 'PC:' + findKey
    pcrd = ''
    mbFindKey = 'MB:' + findKey
    mbrd = ''
    if rankKey is not None and pcFindKey in rankKey.keys():
        pcrd = str(rankKey[pcFindKey])
    if rankKey is not None and mbFindKey in rankKey.keys():
        mbrd = str(rankKey[mbFindKey])
    strTxtPrint = findKey + '( ' + rd + ')' + '          ( PC:' + pcrd + ')' + '     ( MB:' + mbrd + ')'
    return strTxtPrint

def get_rank_common(sIdx, eIdx, findKeyArr, rankKey, rankKeyR, db):
    for findKey in findKeyArr:
        strTxtPrint = set_make_title(rankKey, rankKeyR, findKey)
        print_find_text(str(strTxtPrint))
        for pg in range(sIdx, eIdx + 1):
            find_page(strTxtPrint, findKey, pg, db)

        print('')

def get_rank_pwlink(rowsKey, rankKey, rankKeyR, db):
    findKeyCnt = 0
    for findKeyArr in rowsKey:
        pc_mb = findKeyArr[0]
        findKey = findKeyArr[1]
        view = findKeyArr[2]
        click = findKeyArr[3]
        cost = findKeyArr[4]
        total_cost = findKeyArr[5]

        findKeyCnt += 1
        strTxtPrint = pc_mb.upper() + ':' + str(findKeyCnt) + '. ' + set_make_title(rankKey, rankKeyR, findKey)
        print_find_text(str(strTxtPrint))

        idx = 0
        findFlag = 'N'
        input_data = {}
        if pc_mb == 'pc':
            soup = get_soup_pwlinkMainPC(findKey)
            for tag in soup.select('li'):
                tagTit = tag.find(class_='lnk_tit')
                tagdsc = tag.find(class_='ad_dsc')
                if tagTit:
                    id = tagTit.parent.parent.parent.parent.attrs['id']
                    if id == 'power_link_body' or id == 'biz_site_body':
                        idx += 1
                        if tagTit.attrs['onclick'].find('www.flybeach.co.kr') > -1:
                            pStr = 'INDEX:' + println(str(idx), 10)
                            pStr += println(tagTit.contents[0], 60)
                            tagdscTxt = ''
                            if tagdsc is not None:
                                tagdscTxt = tagdsc.text
                            pStr += println(tagdscTxt, 90)
                            site = tagTit.attrs['onclick']
                            pStr += site[site.find('urlencode'):]
                            print(pStr)
                            findFlag = 'Y'
                            input_data = {'find_key': findKey, 'main_sub': 'main', 'idx': idx, 'idx_total':idx
                                          , 'view': view, 'click': click, 'cost': cost, 'total_cost': total_cost
                                , 'item': tagTit.contents[0], 'item_desc': tagdscTxt, 'pc_mb': 'pc'}

            if findFlag == 'N':
                get_rank_pwlink_sub(findKey, 1, db, idx, view, click, cost, total_cost)
            else:
                input_data['idx_total'] = idx
                insert_history_pwlink(db, input_data)
        else:
            soup = get_soup_pwlinkMainMB(findKey)
            for tag in soup.select('li'):
                tagTxt = tag.text.replace('\n','$').replace('$$','$').replace('$$','$').replace('$$','$').replace('$$','$')
                if tagTxt.find('모바일') > -1:
                    idx += 1
                    if tagTxt.find('flybeach.co.kr') > -1:
                        pStr = 'INDEX:' + println(str(idx), 10)
                        pStr += println(tagTxt,100)
                        print(pStr)
                        findFlag = 'Y'
                        input_data = {'find_key': findKey, 'main_sub': 'main', 'idx': idx
                            , 'view': view, 'click': click, 'cost': cost, 'total_cost': total_cost
                            , 'item': tagTxt.split('$')[1], 'item_desc': tagTxt, 'pc_mb': 'mb'}
            if findFlag == 'Y':
                input_data['idx_total'] = idx
                insert_history_pwlink(db, input_data)

def get_rank_pwlink_sub(findKey, pg, db, idx_total, view, click, cost, total_cost):
    pStr = ' - Sub Page:' + str(pg) + '............................................................................'
    print(pStr)
    soup = get_soup_pwlinkSub(findKey, pg)

    idx = 1
    for tag in soup.select('li'):
        tagTit = tag.find(class_='lnk_tit')
        tagUrl = tag.find(class_='url')
        tagdsc = tag.find(class_='ad_dsc')
        if tagTit:
            if tagUrl.text.find('www.flybeach.co.kr') > -1:
                pStr = println('  '+tag.find(class_='no').text, 5)
                pStr += println(' '+tagTit.text, 60)
                if tagdsc.text is not None:
                    pStr += println(tagdsc.text, 90)
                pStr += println(tagUrl.text, 90)
                input_data = {'find_key': findKey, 'main_sub': 'sub', 'idx': tag.find(class_='no').text, 'idx_total':str(idx_total)
                    , 'view': view, 'click': click, 'cost': cost, 'total_cost': total_cost
                    , 'item': tagTit.contents[0], 'item_desc': tagdsc.text, 'pc_mb':'pc'}
                insert_history_pwlink(db, input_data)
                print(pStr)
                return
        idx += 1

    if pg >= 2:
        return
    else:
        pg += 1
        get_rank_pwlink_sub(findKey, pg, db, idx_total, view, click, cost, total_cost)

def get_rank_pwlink_mb(rowsKey, rankKey, rankKeyR, db):
    # Chrome의 경우 | 아까 받은 chromedriver의 위치를 지정해준다.
    driver = webdriver.Chrome(data_path + "\chromedriver")

    driver.implicitly_wait(3)

    driver.get('https://m.ad.search.naver.com/search.naver?where=m_expd&query=')  # url에 접근한다.

    findKeyCnt = 0
    for findKeyArr in rowsKey:
        pc_mb = findKeyArr[0]
        findKey = findKeyArr[1]
        view = findKeyArr[2]
        click = findKeyArr[3]
        cost = findKeyArr[4]
        total_cost = findKeyArr[5]
        # findKey = '망사비치웨어'

        findKeyCnt += 1
        strTxtPrint = pc_mb.upper() + ':' + str(findKeyCnt) + '. ' + set_make_title(rankKey, rankKeyR, findKey)
        print_find_text(str(strTxtPrint))

        idx = 0
        findFlag = 'N'
        input_data = {}

        # 텍스트 입력
        driver.find_element_by_xpath('//*[@id="query"]').clear()
        driver.find_element_by_xpath('//*[@id="query"]').send_keys(findKey)

        # 버튼을 눌러주자.
        driver.find_element_by_xpath('//*[@id="search"]/button').click()

        viewIdx = 0
        while True:
            viewIdx += 1
            if viewIdx > 10:
                break
            try:
                element = driver.find_element_by_xpath('//*[@id="_get_more"]')
                element.click()
            except:
                break

        html = driver.page_source
        soup = BeautifulSoup(html, 'lxml')

        for tag in soup.select('li'):
            tagTxt = tag.text.replace('\n', '$').replace('$$', '$').replace('$$', '$').replace('$$', '$').replace('$$', '$')
            if tagTxt.find('모바일') > -1:
                idx += 1
                if tagTxt.find('flybeach.co.kr') > -1:
                    pStr = 'INDEX:' + println(str(idx), 10)
                    pStr += println(tagTxt, 100)
                    print(pStr)
                    findFlag = 'Y'
                    input_data = {'find_key': findKey, 'main_sub': 'main', 'idx': idx, 'idx_total': idx*viewIdx
                        , 'view': view, 'click': click, 'cost': cost, 'total_cost': total_cost
                        , 'item': tagTxt.split('$')[1], 'item_desc': tagTxt, 'pc_mb': 'mb'}
                    insert_history_pwlink(db, input_data)
                    break



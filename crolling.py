import os
import openpyxl
pagePrintCnt = 50
prj_path = os.path.dirname(os.path.abspath(__file__))
def make_dir(dir):
    if not os.path.exists(dir):
        os.makedirs(dir)
    return dir

down_path = 'C:\\Users\\chan\\Downloads'
common_path = prj_path+'/common'
data_path = prj_path+'/data'
keyword_pc_path = make_dir(data_path + '/keywordPC')
keyword_mb_path = make_dir(data_path + '/keywordMB')
keyword_r_path = make_dir(data_path + '/keywordR')
key_value_file = common_path + '/' + 'key_value.py'
if not os.path.exists(key_value_file):
    with open(key_value_file, 'w') as f:
        f.write("host='0.0.0.0'\n")
        f.write("port=3307\n")
        f.write("user='hee'\n")
        f.write("password=''\n")
        f.write("db=''\n")
        f.write("charset='utf8mb4'\n")

class searchArrList:
    dict = {"list":[]}

def add_key(keyArrArr):
    kMain = []
    for keyArr in keyArrArr:
        for key in keyArr:
            kMain.append(key)
    return kMain

def get_rank_keywordR():
    rankData = {}
    rank_list = sorted(os.listdir(keyword_r_path), reverse=True)
    for dir in rank_list:
        file = keyword_r_path + '/' + dir
        if file.find('xlsx') > 0 and file.find('연관키워드') > 0:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for row in ws.rows:
                name = row[0].value
                pc = row[1].value
                mb = row[2].value
                if name in rankData:
                    continue
                rankData[name] = {'pc': pc, 'mb': mb}
    return rankData

def get_rank_keyword(pc_mb):
    rankData = {}
    if pc_mb == 'pc':
        keyword_path = keyword_pc_path
    else:
        keyword_path = keyword_mb_path
    rank_list = sorted(os.listdir(keyword_path), reverse=True)
    for dir in rank_list:
        file = keyword_path + '/' + dir
        if file.find('xlsx') > 0 and file.find('키워드 목록') > 0 and file.find('~$키워드 목록') == -1:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for row in ws.rows:
                id = row[0].value
                status = row[1].value
                find_key = row[2].value
                cost = row[4].value
                view = row[6].value
                click = row[7].value
                total_cost = row[11].value
                if id == '' or id == '키워드 ID' or find_key in rankData:
                    continue
                if status != '노출가능':
                    continue
                rankData[find_key] = {'status': status, 'cost': cost, 'view': view, 'click': click, 'total_cost': total_cost}
    return rankData

# rankData = get_rank_key_count()

# def get_rank_key_pwlink():
#     pwLinkDataPc = []
#     pwLinkDataMb = []
#     rank_list = sorted(os.listdir(data_path), reverse=True)
#     for dir in rank_list:
#         file = data_path + '/' + dir
#         if file.find('pwLinkData.xlsx') > 0:
#             wb = openpyxl.load_workbook(file)
#             ws = wb.active
#             for row in ws.rows:
#                 pc_mb = row[0].value
#                 name = row[1].value
#                 if pc_mb == '플라이비치':
#                     if name in pwLinkDataPc:
#                         continue
#                     pwLinkDataPc.append(name)
#                 else:
#                     if name in pwLinkDataMb:
#                         continue
#                     pwLinkDataMb.append(name)
#     return pwLinkDataPc, pwLinkDataMb
########################################################################################################################
# 스포츠/레저 > 수영 > 여성수영복 > 비키니
VK = ['왕뽕비키니', '하이웨스트비키니', '비키니']
VK1 = ['수영복', '여성수영복', '여자수영복']
VK2 = ['원피스비키니', '비키니쇼핑몰', '비키니커버업', '수영복쇼핑몰', '수영복브랜드', '프릴비키니']
VK3 = ['섹시비키니', '비키니브랜드', '비키니수영복', '비키니추천', '심플모노키니', '여자비키니', '여성비키니']

# 스포츠/레저 > 수영 > 여성수영복 > 원피스수영복
VK_OP = ['모노키니', '원피스수영복', '실내수영장수영복', '여자실내수영복', '실내수영복', '여성실내수영복']

# 스포츠/레저 > 수영 > 비치웨어 > 원피스
OP = ['비치웨어', '비치원피스', '비치웨어쇼핑몰', '비치웨어브랜드']

# 스포츠/레저 > 수영 > 비치웨어 > 커플비치웨어
BW_CP = ['커플래쉬가드', '커플레쉬가드', '커플래시가드', '커플레시가드']
BW_CP1 = ['커플래쉬가드세트', '커플수영복', '커플비치웨어', '커플집업래쉬가드']
BW_CP2 = ['신혼여행커플룩']

# 스포츠/레저 > 수영 > 비치웨어 > 상의
BW_TOP = ['래쉬가드', '레쉬가드', '래시가드', '레시가드']
BW_TOP1 = ['래쉬가드추천', '수영복래쉬가드', '래쉬가드브랜드', '여성집업래쉬가드', '수영복래쉬가드', '루즈핏래쉬가드', '크롭래쉬가드']
BW_TOP2 = ['여성래쉬가드', '여성레쉬가드', '여성래시가드', '여성레시가드']
BW_TOP3 = ['여자래쉬가드', '여자레쉬가드', '여자래시가드', '여자레시가드']
BW_TOP4 = ['집업래쉬가드', '집업레쉬가드', '집업래시가드', '집업레시가드']
BW_TOP5 = ['남자래쉬가드', '남자레쉬가드', '남자래시가드', '남자레시가드']
BW_TOP6 = ['남성래쉬가드', '남성레쉬가드', '남성래시가드', '남성레시가드']

# 스포츠/레저 > 수영 > 비치웨어 > 상하세트
# BW_TOP_SET = ['래쉬가드세트', '여성래쉬가드세트']
########################################################################################################################
VK = add_key([VK, VK1, VK2, VK3])
VK_OP = add_key([VK_OP])
OP = add_key([OP])
BW_CP = add_key([BW_CP, BW_CP1, BW_CP2])
BW_TOP = add_key([BW_TOP, BW_TOP1, BW_TOP2, BW_TOP3, BW_TOP4, BW_TOP5, BW_TOP6])

key = add_key([VK, VK_OP, OP, BW_CP, BW_TOP])
########################################################################################################################
# Sum
keySum = [
 {'id':'BW_TOP', 'key':BW_TOP      , 'name':'래쉬가드추천 FB1168_R52'                     , 'mid1':'14705104256'}
,{'id':'BW_TOP', 'key':BW_TOP      , 'name':'래쉬가드추천 ( 여성레쉬가드 ) FB1168_R52'    , 'mid1':'9954709419'}
,{'id':'BW_TOP', 'key':BW_TOP      , 'name':'여성레쉬가드 래쉬가드 FB1168_R55'            , 'mid1':'80656672132'}

,{'id':'BW_CP' , 'key':BW_CP       , 'name':'커플레쉬가드 커플수영복 FB1168_R54'          , 'mid1':'80651469478'}
,{'id':'BW_CP' , 'key':BW_CP       , 'name':'신혼여행커플룩 허니문커플룩'                 , 'mid1':'11751841925'}
,{'id':'BW_CP' , 'key':BW_CP       , 'name':'[플라이비치] 신혼여행커플룩 허니문커플룩'    , 'mid1':'10114258285'}
,{'id':'BW_CP' , 'key':BW_CP       , 'name':'[플라이비치]신혼여행커플룩,허니문웨어'       , 'mid1':'12727818144'}

,{'id':'VK'    , 'key':VK          , 'name':'왕뽕 하이웨스트 비키니 모노키니 FB1168_R57'  , 'mid1':'10054140280'}

,{'id':'VK_OP' , 'key':VK_OP       , 'name':'왕뽕 원피스수영복 모노키니 FB1168_R56'       , 'mid1':'80656785020'}

,{'id':'OP'    , 'key':OP          , 'name':'비치원피스 비치웨어 FB1168_R58'              , 'mid1':'80657855013'}
,{'id':'OP'    , 'key':OP          , 'name':'[플라이비치] 비치원피스 여름원피스'          , 'mid1':'10926541260'}

]
########################################################################################################################
# Naver
# {'key':'커플레쉬가드'       , 'name':'커플레쉬가드 커플수영복 FB1168_R54'  , 'mid1':'80651469478', 'mid2':'', 'cnt':'1'}
# ,{'key':'커플수영복'         , 'name':'커플레쉬가드 커플수영복 FB1168_R54'  , 'mid1':'80651469478', 'mid2':'', 'cnt':'1'}
# ,{'key':'비치원피스'         , 'name':'비치원피스 비치웨어 FB1168_R58'      , 'mid1':'80657855013', 'mid2':'', 'cnt':'1'}
keyNv = [
 {'key':'커플래쉬가드'       , 'name':'비치서핑 래쉬가드'                   , 'mid1':'14638288265', 'mid2':'80452092166', 'cnt':'2'}
,{'key':'래쉬가드'           , 'name':'서프알로하 나시래쉬가드'             , 'mid1':'14489384918', 'mid2':'80495094273', 'cnt':'3'}
,{'key':'커플레쉬가드'       , 'name':'서프알로하 래쉬가드(커플상의) '      , 'mid1':'14420533996', 'mid2':'', 'cnt':'8'}
,{'key':'래쉬가드'           , 'name':'래쉬가드추천 FB1168_R52'             , 'mid1':'14705104256', 'mid2':'9954709419', 'cnt':'5'}
,{'key':'왕뽕비키니'         , 'name':'왕뽕 하이웨스트 비키니 모노키니'     , 'mid1':'10591394440', 'mid2':'10054140280', 'cnt':'3'}
,{'key':'커플수영복'         , 'name':'블루에스닉 커플수영복'               , 'mid1':'14595351003', 'mid2':'', 'cnt':'6'}
,{'key':'모노키니'           , 'name':'크로스 모노키니'                     , 'mid1':'14355364202', 'mid2':'80418993407', 'cnt':'6'}
,{'key':'신혼여행커플룩'     , 'name':'신혼여행커플룩 허니문웨어'           , 'mid1':'12727818144', 'mid2':'11751841925', 'cnt':'1'}
,{'key':'모노키니'           , 'name':'섹시오프 모노키니'                   , 'mid1':'14616533726', 'mid2':'80494394781', 'cnt':'17'}
,{'key':'비치원피스'         , 'name':'플라이비치 비치원피스 여름'          , 'mid1':'10926541260', 'mid2':'80657855013', 'cnt':'6'}
]
########################################################################################################################



